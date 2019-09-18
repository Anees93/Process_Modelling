from Process_Modelling.CustomerClass import Customer
from Process_Modelling.StationClass import Station
from openpyxl import load_workbook
from random import random
import numpy as np
from scipy.stats import norm


def print_results(Customers, lTotalTicks):
    print("ID:  ", "Entered@    ", "Left@   ",
          "@station:   ", "Idle: ", "IdleTime:",
          "TimeSpent: ", "%Idle:")

    for customer in Customers:
        if customer.get_Left() is not None:
            time_in_system = customer.get_Left() - customer.get_Entered() + 1
        else:
            time_in_system = lTotalTicks - customer.get_Entered() + 1

        print(customer.get_CustID(), "      ", customer.get_Entered(), "        ", customer.get_Left(),
              "     ", customer.get_Station(), "        ", customer.get_IsIdle(), "     ",
              customer.get_IdleTime(), "        ", np.round(time_in_system, decimals=2),
              "     ", np.round(customer.get_IdleTime() * 100 / time_in_system, decimals=2))


def main():
    Customers = []
    Stations = []
    wb = load_workbook(filename="Data.xlsm")
    ws = wb.worksheets[0]
    lNmbrStas = ws.cell(row=2, column=3).value

    for i in range(0, lNmbrStas):
        Stations.append(Station(i+1))
        Stations[i].set_StaMean(ws.cell(row=i + 5, column=3).value)
        Stations[i].set_StaSD(ws.cell(row=i + 5, column=4).value)
        Stations[i].set_NextSta(ws.cell(row=i + 5, column=5).value)
        Stations[i].set_StaIsIdle(1)

    # Total ticks for each run
    lTotalTicks = 2880  # each tick represents 10 seconds in an eight-hour day
    lCreateNextAt = 1
    # Set CustID counter to 1
    lCustIDCntr = 1
    # Prepare customer arrival data
    N = 54
    mean = 20

    print("Simulation Started")
    # Outer For Next loop runs for the total ticks set above
    for i in range(1, lTotalTicks+1):
        # print("START OF TICKER", i)
        # Check if a customer is due to be created on this tick, if not, do nothing.
        if i == lCreateNextAt:
            # print("Customer#", lCustIDCntr, " Arrival in Ticker", i)
            Customers.append(Customer(lCustIDCntr))
            Customers[lCustIDCntr - 1].set_StartTime(1)
            Customers[lCustIDCntr - 1].set_NextSta(1)
            Customers[lCustIDCntr - 1].set_IsIdle(1)
            Customers[lCustIDCntr - 1].set_IdleTime(0)
            Customers[lCustIDCntr - 1].set_Entered(i)
            lCreateNextAt += int(np.random.choice(np.random.poisson(mean, size=N)))
            lCustIDCntr += 1

        # Start checking each customer's status. If they are idle,
        # check which station they are in. If it's -1, they are done.
        # If it's any other station, generate a processing time and
        # set the customer's IsIdle property to 0.

        for customer in Customers:
            if customer.get_Station() != -1:
                if customer.get_IsIdle() == 1:
                    # If the customer remains idle for this tick, add 1 to idle time.
                    if Stations[customer.get_NextSta() - 1].get_StaIsIdle() is 0:
                        customer.set_IdleTime(1 + customer.get_IdleTime())
                        # print("Customer#", customer.get_CustID(), " Waited another ticker")
                    # If not, set as not idle and generate processing time.
                    else:
                        customer.set_Station(customer.get_NextSta())
                        Stations[customer.get_Station()-1].set_StaIsIdle(0)
                        customer.set_NextSta(Stations[customer.get_Station()-1].get_NextSta())
                        endTime = i + norm.ppf(random(), Stations[customer.get_Station()-1].get_StaMean(),
                                               Stations[customer.get_Station()-1].get_StaSD())
                        customer.set_EndTime(int(endTime))
                        # print("Customer#", customer.get_CustID(),"Seized Station#", customer.get_Station(),
                        #       "until ", int(endTime), "Ticker")
                        # Make sure the customer spends at least one tick in a station
                        if customer.get_EndTime() <= i:
                            customer.set_EndTime(i+1)
                        customer.set_IsIdle(0)
                        customer.set_StartTime(i)

                # If a customer's EndTime = the current tick, update Idle status to true
                # and move it to the next station
                elif customer.get_EndTime() == i:
                    Stations[customer.get_Station()-1].set_StaIsIdle(1)
                    customer.set_IsIdle(1)
                    # print("Customer#", customer.get_CustID(), "Released Station#", customer.get_Station())
                    customer.set_Station(customer.get_NextSta())

                    # If the customer's station is -1, it is done.
                    # Set station to -1 and record the time it left the system.
                    if customer.get_Station() == -1:
                        customer.set_Left(i)
                        # print("Customer#", customer.get_CustID(), "has left the system")

            else:
                # Sets last station to idle when a customer leaves it
                Stations[lNmbrStas-1].set_StaIsIdle(1)

    print_results(Customers, lTotalTicks)


if __name__ == "__main__":
    main()


